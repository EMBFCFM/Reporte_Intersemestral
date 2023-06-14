import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

# Cargamos el archivo excel existente
archivo_existente = r"C:\Users\Usuario\Desktop\PRUEBA DE HORARIOS\Reporte intersemestrales 2023.xlsx"
wb_existente = openpyxl.load_workbook(archivo_existente)
hoja_existente = wb_existente.active

# Creamos un nuevo archivo excel
wb_nuevo = Workbook()
hoja_nuevo = wb_nuevo.active

# Agregamos encabezados de columna
datos_columna = ['ESTATUS', 'CVE. MAT.', 'MATERIA', 'Frec. Mat.', 'NIVEL DE GRUPO', 'MODALIDAD', 'PLAN', 'NO. EMPLEADO',
                 'MAESTRO', 'HORARIO', 'FREC', 'GPO', 'AULA']

for i, dato in enumerate(datos_columna, start=1):
    celda = hoja_nuevo.cell(row=8, column=i)
    celda.value = dato

# Aquí comienza la búsqueda
columna_busqueda = "A"
nombres_busqueda = ["Equivalencia"]

filas_encontradas = []
columna_index = column_index_from_string(columna_busqueda)

for nombre_busqueda in nombres_busqueda:
    for fila_index, fila in enumerate(hoja_existente.iter_rows(min_row=1, max_row=hoja_existente.max_row,
                                                               min_col=columna_index, max_col=columna_index), start=1):
        celda = fila[0]
        if celda.value == nombre_busqueda:
            filas_encontradas.append((fila_index, fila))

# Imprimir las filas encontradas
if filas_encontradas:
    for nombre_busqueda in nombres_busqueda:
        print("Resultados para:", nombre_busqueda)
        resultados = [(fila_index, fila) for fila_index, fila in filas_encontradas if fila[0].value == nombre_busqueda]
        if resultados:
            for fila_index, fila in resultados:
                materia = fila[column_index_from_string("C") - 1].value if len(fila) >= column_index_from_string("C") else ""
                print("Fila:", fila_index)
                print("Materia:", materia)
                for celda in fila:
                    print(celda.value, end=" ")
                print()
        else:
            print("Nombre no encontrado.")
        print()
else:
    print("No se encontraron coincidencias.")

# Guardar el archivo
archivo_nuevo = "Reporte.xlsx"
wb_nuevo.save(archivo_nuevo)
